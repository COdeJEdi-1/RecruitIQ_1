"""
Interview Ready — AI voice-screening review page.
Reads call answers from the OmniDimension-linked Google Sheet and shows
recruiters who's ready for interview scheduling, per utils/voice_screening.py.
"""

import io
from datetime import datetime

from flask import Blueprint, render_template, send_file

from routes.auth import login_required
from utils.voice_screening import (
    fetch_and_score,
    SENTIMENT_RATING_LABELS,
    COMMUNICATION_RATING_LABELS,
)

interview_ready_bp = Blueprint("interview_ready", __name__)

# Incomplete (no-answer / too-short) calls carry no sentiment or communication
# signal and can vastly outnumber real, analysable calls — cap how many clutter
# the table while the stat tile above still reports the true total.
MAX_INCOMPLETE_ROWS_SHOWN = 5


def _compute_analytics(scored):
    """Sentiment/communication distribution + average score, over candidates
    that actually have a verdict (i.e. excludes Incomplete)."""
    sentiment_dist = {rating: 0 for rating in sorted(SENTIMENT_RATING_LABELS, reverse=True)}
    communication_dist = {rating: 0 for rating in sorted(COMMUNICATION_RATING_LABELS, reverse=True)}

    scores = [c["score"] for c in scored if c.get("score") is not None]
    for c in scored:
        if c.get("sentiment_rating") in sentiment_dist:
            sentiment_dist[c["sentiment_rating"]] += 1
        if c.get("communication_rating") in communication_dist:
            communication_dist[c["communication_rating"]] += 1

    return {
        "sentiment_dist": [
            {"rating": r, "label": SENTIMENT_RATING_LABELS[r], "count": sentiment_dist[r]}
            for r in sentiment_dist
        ],
        "communication_dist": [
            {"rating": r, "label": COMMUNICATION_RATING_LABELS[r], "count": communication_dist[r]}
            for r in communication_dist
        ],
        "avg_score": round(sum(scores) / len(scores), 1) if scores else None,
        "analysed_count": len(scores),
    }


@interview_ready_bp.route("/interview-ready")
def index():
    candidates, error = fetch_and_score()

    counts = {
        "Ready for Interview": 0,
        "Needs Manual Call": 0,
        "Manual Review": 0,
        "Not Suitable": 0,
        "Incomplete": 0,
    }
    for c in candidates:
        counts[c["verdict"]] = counts.get(c["verdict"], 0) + 1

    scored = [c for c in candidates if c["verdict"] != "Incomplete"]
    incomplete = [c for c in candidates if c["verdict"] == "Incomplete"]
    hidden_incomplete_count = max(0, len(incomplete) - MAX_INCOMPLETE_ROWS_SHOWN)
    candidates_shown = scored + incomplete[-MAX_INCOMPLETE_ROWS_SHOWN:]

    analysed = [c for c in scored if c["verdict"] != "Needs Manual Call"]
    analytics = _compute_analytics(analysed)
    roles = sorted({c["role"] for c in candidates_shown if c.get("role")})

    return render_template(
        "interview_ready.html",
        candidates=candidates_shown,
        counts=counts,
        error=error,
        hidden_incomplete_count=hidden_incomplete_count,
        analytics=analytics,
        roles=roles,
    )


@interview_ready_bp.route("/api/interview-ready-export")
@login_required
def export():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    candidates, error = fetch_and_score()
    if error:
        return {"error": error}, 502

    scored = [c for c in candidates if c["verdict"] != "Incomplete"]
    _order = {
        "Needs Manual Call": 0,
        "Ready for Interview": 1,
        "Manual Review": 2,
        "Not Suitable": 3,
        "Incomplete": 4,
    }
    scored.sort(key=lambda c: (_order.get(c["verdict"], 9), c["score"] is None, -(c["score"] or 0)))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Interview Ready"

    headers = [
        "Name", "Role", "Phone", "Verdict", "Recording Consent", "Score", "Experience (yrs)",
        "Current CTC", "Expected CTC", "Notice / Joining", "Location",
        "Sentiment", "Sentiment Rating (1-5)", "Communication", "Communication Rating (1-5)",
    ]
    hdr_fill = PatternFill("solid", fgColor="830026")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="E0E0E0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_idx, c in enumerate(scored, start=2):
        values = [
            c.get("name", ""), c.get("role", ""), c.get("phone", ""), c.get("verdict", ""),
            c.get("recording_consent") or "—", c.get("score"),
            c.get("years_experience"), c.get("current_ctc", ""), c.get("expected_ctc", ""),
            c.get("notice_period", ""), c.get("current_location", ""),
            c.get("computed_sentiment") or "Not enough data", c.get("sentiment_rating"),
            c.get("communication_label") or "Not enough data", c.get("communication_rating"),
        ]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    widths = [20, 20, 16, 20, 16, 8, 14, 14, 14, 16, 16, 14, 12, 24, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"interview_ready_{ts}.xlsx",
    )
