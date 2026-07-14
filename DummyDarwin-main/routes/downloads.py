"""
Download request workflow — recruiter requests bulk resume download,
admin approves/rejects, recruiter gets notified and can then download.
"""

import io
import json
import sys
import uuid
import zipfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from flask import Blueprint, jsonify, request, send_file, session

from routes.auth import login_required, is_admin, current_user_id

downloads_bp = Blueprint("downloads", __name__)

_JD_PROTOTYPE_DIR = Path(__file__).parent.parent.parent / "jd_prototype"
_DARWIN_DATA  = _JD_PROTOTYPE_DIR / "darwin_data"
_REQUESTS_FILE = _DARWIN_DATA / "download_requests.json"
_CANDIDATES_FILE = _DARWIN_DATA / "candidates.json"
_SCORES_FILE   = _DARWIN_DATA / "candidate_scores.json"
_USERS_FILE    = _DARWIN_DATA / "users.json"
_RESUME_BASE   = _JD_PROTOTYPE_DIR / "static" / "uploads" / "candidates"

if str(_JD_PROTOTYPE_DIR) not in sys.path:
    sys.path.insert(0, str(_JD_PROTOTYPE_DIR))

from scoring_service import AUTO_CALL_SCORE_THRESHOLD  # noqa: E402


def _load(path: Path) -> list:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return []


def _save(path: Path, data: list):
    path.parent.mkdir(exist_ok=True)
    path.write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")


def _load_requests() -> list:
    if not _REQUESTS_FILE.exists():
        return []
    return _load(_REQUESTS_FILE)


def _save_requests(data: list):
    _save(_REQUESTS_FILE, data)


def _user_name(user_id: str) -> str:
    for u in _load(_USERS_FILE):
        if u["user_id"] == user_id:
            return u.get("full_name") or u.get("username", user_id[:8])
    return user_id[:8]


def _apply_filters(candidates: list, scores: dict, filters: dict) -> list:
    date_from = filters.get("date_from") or ""
    date_to   = filters.get("date_to")   or ""
    min_score = float(filters.get("min_score") or 0)
    platforms = filters.get("platforms") or []
    job_ids   = filters.get("job_ids")   or []

    result = []
    for c in candidates:
        applied = c.get("applied_at", "")[:10]
        if date_from and applied < date_from:
            continue
        if date_to and applied > date_to:
            continue
        if platforms and c.get("platform_source") not in platforms:
            continue
        if job_ids and c.get("darwinbox_job_id") not in job_ids:
            continue
        sc = scores.get(c["candidate_id"])
        overall = sc["overall_score"] if sc else 0
        if overall < min_score:
            continue
        result.append(c)
    return result


# ── Recruiter: submit download request ───────────────────────────────────────

@downloads_bp.route("/api/download-request", methods=["POST"])
@login_required
def submit_request():
    if is_admin():
        return jsonify({"error": "Admins can download directly."}), 400

    data    = request.get_json(silent=True) or {}
    filters = {
        "date_from": data.get("date_from", ""),
        "date_to":   data.get("date_to",   ""),
        "min_score": data.get("min_score", 0),
        "platforms": data.get("platforms", []),
        "job_ids":   data.get("job_ids",   []),
    }

    uid = current_user_id()
    darwin_jobs = _load(_DARWIN_DATA / "darwinbox_jobs.json")
    owned_ids   = {j["darwinbox_job_id"] for j in darwin_jobs if j.get("created_by") == uid}
    candidates  = [c for c in _load(_CANDIDATES_FILE) if c.get("darwinbox_job_id") in owned_ids]
    scores      = {s["candidate_id"]: s for s in _load(_SCORES_FILE)}
    matched     = _apply_filters(candidates, scores, filters)

    req_id  = str(uuid.uuid4())
    new_req = {
        "request_id":    req_id,
        "requester_id":  uid,
        "requester_name": _user_name(uid),
        "filters":       filters,
        "match_count":   len(matched),
        "status":        "pending",
        "created_at":    datetime.now().isoformat(),
        "reviewed_at":   None,
        "reviewed_by":   None,
        "reviewer_name": None,
        "reject_reason": None,
    }
    reqs = _load_requests()
    reqs.append(new_req)
    _save_requests(reqs)
    return jsonify({"success": True, "request_id": req_id, "match_count": len(matched)})


# ── Recruiter: poll status ────────────────────────────────────────────────────

@downloads_bp.route("/api/download-request/<req_id>/status")
@login_required
def request_status(req_id):
    uid = current_user_id()
    for r in _load_requests():
        if r["request_id"] == req_id:
            if not is_admin() and r["requester_id"] != uid:
                return jsonify({"error": "Forbidden"}), 403
            return jsonify(r)
    return jsonify({"error": "Not found"}), 404


# ── Admin: list all requests ──────────────────────────────────────────────────

@downloads_bp.route("/api/download-requests")
@login_required
def list_requests():
    if not is_admin():
        return jsonify({"error": "Admin only"}), 403
    reqs = sorted(_load_requests(), key=lambda r: r["created_at"], reverse=True)
    return jsonify(reqs)


# ── Admin: approve or reject ──────────────────────────────────────────────────

@downloads_bp.route("/api/download-request/<req_id>/review", methods=["POST"])
@login_required
def review_request(req_id):
    if not is_admin():
        return jsonify({"error": "Admin only"}), 403

    data   = request.get_json(silent=True) or {}
    action = data.get("action")
    if action not in ("approve", "reject"):
        return jsonify({"error": "action must be approve or reject"}), 400

    reqs = _load_requests()
    for r in reqs:
        if r["request_id"] == req_id:
            if r["status"] != "pending":
                return jsonify({"error": "Already reviewed"}), 409
            r["status"]        = "approved" if action == "approve" else "rejected"
            r["reviewed_at"]   = datetime.now().isoformat()
            r["reviewed_by"]   = current_user_id()
            r["reviewer_name"] = _user_name(current_user_id())
            r["reject_reason"] = data.get("reason", "")
            _save_requests(reqs)
            return jsonify({"success": True, "status": r["status"]})

    return jsonify({"error": "Not found"}), 404


# ── Recruiter: download approved ZIP ─────────────────────────────────────────

@downloads_bp.route("/api/download-request/<req_id>/download")
@login_required
def download_bundle(req_id):
    uid    = current_user_id()
    target = next((r for r in _load_requests() if r["request_id"] == req_id), None)

    if not target:
        return jsonify({"error": "Request not found"}), 404
    if not is_admin() and target["requester_id"] != uid:
        return jsonify({"error": "Forbidden"}), 403
    if target["status"] != "approved":
        return jsonify({"error": "Not yet approved"}), 403

    darwin_jobs = _load(_DARWIN_DATA / "darwinbox_jobs.json")
    all_cands   = _load(_CANDIDATES_FILE)
    if not is_admin():
        owned_ids = {j["darwinbox_job_id"] for j in darwin_jobs if j.get("created_by") == uid}
        all_cands = [c for c in all_cands if c.get("darwinbox_job_id") in owned_ids]

    scores  = {s["candidate_id"]: s for s in _load(_SCORES_FILE)}
    matched = _apply_filters(all_cands, scores, target["filters"])
    job_map = {j["darwinbox_job_id"]: j["role_title"] for j in darwin_jobs}

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        lines = ["Name,Email,Phone,Role,Platform,Score,Applied\n"]
        for c in matched:
            sc      = scores.get(c["candidate_id"])
            overall = sc["overall_score"] if sc else ""
            lines.append(
                f"{c['name']},{c['email']},{c.get('phone','')},{job_map.get(c['darwinbox_job_id'],'')},{c['platform_source']},{overall},{c['applied_at'][:10]}\n"
            )
        zf.writestr("candidates_manifest.csv", "".join(lines))

        for c in matched:
            rf = c.get("resume_file", "")
            if rf:
                resume_path = _RESUME_BASE / Path(rf).name
                if resume_path.exists():
                    safe = f"{c['name'].replace(' ','_')}_{c['candidate_id'][:6]}{resume_path.suffix}"
                    zf.write(resume_path, safe)

    buf.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    return send_file(buf, mimetype="application/zip", as_attachment=True,
                     download_name=f"resumes_{ts}.zip")


# ── Top-3 shortlist Excel export ──────────────────────────────────────────────

@downloads_bp.route("/api/top3-export")
@login_required
def top3_export():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    uid   = current_user_id()
    admin = is_admin()

    darwin_jobs = _load(_DARWIN_DATA / "darwinbox_jobs.json")
    all_cands   = _load(_CANDIDATES_FILE)
    scores      = {s["candidate_id"]: s for s in _load(_SCORES_FILE)}

    if not admin:
        owned_ids = {j["darwinbox_job_id"] for j in darwin_jobs if j.get("created_by") == uid}
        all_cands = [c for c in all_cands if c.get("darwinbox_job_id") in owned_ids]

    job_map = {j["darwinbox_job_id"]: j["role_title"] for j in darwin_jobs}

    by_role = defaultdict(list)
    for c in all_cands:
        sc      = scores.get(c["candidate_id"])
        overall = sc["overall_score"] if sc else 0
        by_role[c.get("darwinbox_job_id", "unknown")].append((overall, c))

    rows = []
    for job_id, items in by_role.items():
        items.sort(key=lambda x: x[0], reverse=True)
        role_title = job_map.get(job_id, "Unknown Role")
        for rank, (score, c) in enumerate(items[:3], start=1):
            rows.append({
                "Role":         role_title,
                "Rank":         f"#{rank}",
                "Name":         c.get("name", ""),
                "Email":        c.get("email", ""),
                "Phone":        c.get("phone", ""),
                "Score":        score,
                "Platform":     c.get("platform_source", "").title(),
                "Applied Date": c.get("applied_at", "")[:10],
            })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Top 3 Candidates"

    headers     = ["Role", "Rank", "Name", "Email", "Phone", "Score", "Platform", "Applied Date"]
    hdr_fill    = PatternFill("solid", fgColor="830026")
    hdr_font    = Font(bold=True, color="FFFFFF", size=11)
    thin        = Side(style="thin", color="E0E0E0")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, start=1):
        cell            = ws.cell(row=1, column=col, value=h)
        cell.fill       = hdr_fill
        cell.font       = hdr_font
        cell.alignment  = Alignment(horizontal="center", vertical="center")
        cell.border     = border

    rank_colors = {"#1": "FFF8E1", "#2": "F3F3F3", "#3": "FFF3E0"}
    for row_idx, row in enumerate(rows, start=2):
        fill = PatternFill("solid", fgColor=rank_colors.get(row["Rank"], "FFFFFF"))
        for col, h in enumerate(headers, start=1):
            cell           = ws.cell(row=row_idx, column=col, value=row[h])
            cell.fill      = fill
            cell.border    = border
            cell.alignment = Alignment(vertical="center")

    for i, w in enumerate([28, 7, 22, 30, 16, 8, 12, 14], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"top3_candidates_{ts}.xlsx",
    )


# ── Qualified (auto-call) candidates Excel export ─────────────────────────────
# Filters on AUTO_CALL_SCORE_THRESHOLD (imported above) — this list is exactly
# who the auto-screening call pipeline will call, by construction.


@downloads_bp.route("/api/qualified-export")
@login_required
def qualified_export():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    uid   = current_user_id()
    admin = is_admin()

    darwin_jobs = _load(_DARWIN_DATA / "darwinbox_jobs.json")
    all_cands   = _load(_CANDIDATES_FILE)
    scores      = {s["candidate_id"]: s for s in _load(_SCORES_FILE)}

    if not admin:
        owned_ids = {j["darwinbox_job_id"] for j in darwin_jobs if j.get("created_by") == uid}
        all_cands = [c for c in all_cands if c.get("darwinbox_job_id") in owned_ids]

    job_map = {j["darwinbox_job_id"]: j["role_title"] for j in darwin_jobs}

    qualified = []
    for c in all_cands:
        sc = scores.get(c["candidate_id"])
        overall = sc["overall_score"] if sc else 0
        if overall < AUTO_CALL_SCORE_THRESHOLD:
            continue
        qualified.append((overall, c))

    qualified.sort(key=lambda x: x[0], reverse=True)

    rows = []
    for score, c in qualified:
        rows.append({
            "Name":  c.get("name", ""),
            "Phone": c.get("phone", ""),
            "Email": c.get("email", ""),
            "Role":  job_map.get(c.get("darwinbox_job_id"), "Unknown Role"),
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Qualified Candidates"

    headers  = ["Name", "Phone", "Email", "Role"]
    hdr_fill = PatternFill("solid", fgColor="830026")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    thin     = Side(style="thin", color="E0E0E0")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, start=1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border

    for row_idx, row in enumerate(rows, start=2):
        for col, h in enumerate(headers, start=1):
            cell           = ws.cell(row=row_idx, column=col, value=row[h])
            cell.border    = border
            cell.alignment = Alignment(vertical="center")

    for i, w in enumerate([22, 16, 30, 22], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"qualified_candidates_{ts}.xlsx",
    )
